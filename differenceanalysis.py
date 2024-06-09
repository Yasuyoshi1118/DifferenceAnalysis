import PySimpleGUI as sg
import os
import csv
import numpy as np
import openpyxl
import time

check_dic = {
    '1':'TVECS No検索'
}

layout = [
    [
    [sg.Text("正解ファイル選択"), sg.InputText(key="path"), sg.FileBrowse(key="selected_file_path"),
    sg.Text("確認列："), sg.InputText("2",key="analysis_column",size=(5,1)),
    sg.Text("表示名称列："), sg.InputText("3",key="name_column",size=(5,1))],
    ],
    [
    [sg.Text("解析ファイル選択"), sg.InputText(key="path2"), sg.FileBrowse(key="selected_file_path2"),
    sg.Text("確認列："), sg.InputText("2",key="analysis_column2",size=(5,1))],
    ],
    [
     [sg.Text("出力CSVファイル名："), sg.InputText("output.xlsx",key="output_xlsx_name")],
    ],
    [
    sg.Checkbox(item[1],key=item[0])for item in check_dic.items()
    ],
   [
    sg.Button('解析＆出力',key='out',size =(15 ,2),button_color =('#ffffff','#000000'))
   ]
]

window = sg.Window("ファイル選択", layout)


def diff_analysis():
    tvecsno_check = 0
    while True:
        event, values = window.read()
        if event =="out":
            #ファイルの取り込み
            for value in values.items():
                #if value[0]==1:
                if value[1]==True:
                    print("==================================")
                    print("TVECS No検索を開始します！！！！！！")
                    print("==================================")
                    tvecsno_check = 1
                else:
                    print("==================")
                    print("通常検索を開始します")
                    print("==================")


            #print(str(values['selected_file_path']))
            file_path = os.path.dirname(str(values['selected_file_path']))
            file_name = os.path.basename(str(values['selected_file_path']))
            #print("file_path : ",file_path)
            #print("log_file_name : ",file_name)
            read_csv(file_path,file_name,1)

            #print(str(values['selected_file_path2']))
            file_path2 = os.path.dirname(str(values['selected_file_path2']))
            file_name2 = os.path.basename(str(values['selected_file_path2']))
            #print("file_path2 : ",file_path2)
            #print("log_file_name2 : ",file_name2)
            read_csv(file_path2,file_name2,2)

            #xlsxの作成
            output_xlsx_name = values["output_xlsx_name"]
            name_ref_sheet = "Sheet1des"
            #print("output_xlsx_name : ",output_xlsx_name)
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = name_ref_sheet
            wb.save(output_xlsx_name)

            analysis_col = values["analysis_column"]
            analysis_col2 = values["analysis_column2"]
            name_col = values["name_column"]
            time.sleep(1)
            difference_analysis(output_xlsx_name,name_ref_sheet,analysis_col,analysis_col2,tvecsno_check,name_col)

#csvログ一次格納リスト
list1 = []
list2 = []

def read_csv(file_path,file_name,list1or2):
    #プログラムがあるフォルダと読み込むファイルパスをマージしパスを生成
    file_link = os.path.join(file_path,file_name)

    if (list1or2==1):
        with open(file_link,"r") as f:
             reader = csv.reader(f)
             for line in reader:
                 list1.append(line)
    elif (list1or2==2):
        with open(file_link,"r") as f:
             reader = csv.reader(f)
             for line in reader:
                 list2.append(line)


def difference_analysis(output_xlsx_name,name_ref_sheet,analysis_col,analysis_col2,tvecsno_check,name_col):

    #xlsxにデータを格納していく
    ref_file = openpyxl.load_workbook(output_xlsx_name)
    ref_sheet = ref_file[name_ref_sheet]

    ref_sheet.cell(row=1,column=2).value="row_num"
    ref_sheet.cell(row=1,column=3).value="項目"
    #格納ループの開始
    xy = 1
    sheet_write = 2
    for i in list1:
        xy = xy + 1
        data_i = i[int(analysis_col)-1]
        if tvecsno_check==1:
            if len(data_i)==5:
                data_i = "0"+ data_i
                print("******************data_i*******************:",data_i)
        check = 0
        for j in list2:
            data_j = j[int(analysis_col2)-1]
            if tvecsno_check==1:
                if len(data_j)==5:
                    data_j = "0"+ data_j
                    print("******************data_j*******************:",data_j)
            print("行；",xy,"   data_j ; ",data_j)
            if(data_i == data_j):
                print("******************あり*******************")
                check = 1
                break
            elif(data_i == ""):
                print("=========================================")
                check = 1
                break
        if(check==0):
            ref_sheet.cell(row=sheet_write,column=2).value=xy-1
            ref_sheet.cell(row=sheet_write,column=3).value=i[int(analysis_col)-1]
            ref_sheet.cell(row=sheet_write,column=4).value=i[int(name_col)-1]
            sheet_write = sheet_write + 1
            #print("check=0")
        print("行；",xy,"   data_i ; ",data_i)
        check = 0

    ref_file.save(output_xlsx_name)#ファイルの保存
    print("ループ終了しました  「 Ctrl + C 」を押して終了してください")




if __name__ == '__main__':
    diff_analysis()
